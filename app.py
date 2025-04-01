import streamlit as st
import pandas as pd
import openpyxl
import locale
from io import BytesIO
import altair as alt
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Configure la locale pour le formatage
locale.setlocale(locale.LC_ALL, '')

# Convertit une colonne en float après avoir nettoyé les caractères non numériques
def column_to_float(dataframe, column):
    dataframe[column] = dataframe[column].astype(str).str.replace('€', '').str.replace(',', '.')
    dataframe[column] = dataframe[column].astype(float)
    return dataframe

# Remplace les noms de mois français par leurs équivalents anglais
def french_to_english(dataframe):
    month_dict = {
        'janv.': 'Jan', 'févr.': 'Feb', 'mars': 'Mar', 'avr.': 'Apr', 'mai': 'May',
        'juin': 'Jun', 'juil.': 'Jul', 'août': 'Aug', 'sept.': 'Sep', 'oct.': 'Oct',
        'nov.': 'Nov', 'déc.': 'Dec'
    }
    for key, value in month_dict.items():
        dataframe['Date'] = dataframe['Date'].str.replace(key, value)
    return dataframe

# Filtre les données en fonction des dates et des descriptions sélectionnées
def filter_data(dataframe, start_date, end_date, description_list):
    return dataframe[
        (dataframe['Date'] >= start_date) & 
        (dataframe['Date'] <= end_date) & 
        (dataframe['Description'].isin(description_list))
    ]

# Formate une valeur en euros avec une différence optionnelle
def metric_in_euros(value, former_value=0):
    return f'{value:,.2f} €', f'{(value - former_value):,.2f} €'

# Calcule les totaux pour les verres et bouteilles
def glass_bottle_total(dataframe, glass_description_list, bottle_description_list):
    glass_df = dataframe[dataframe['Description'].isin(glass_description_list)]
    bottle_df = dataframe[dataframe['Description'].isin(bottle_description_list)]
    glass_total = glass_df['Quantité'].sum()
    bottle_total = bottle_df['Quantité'].sum()
    totals_df = pd.DataFrame({
        'Item': ['Verres', 'Bouteilles'],
        'Quantité': [glass_total, bottle_total],
        'Total (en €)': [glass_df['Prix (TTC)'].sum(), bottle_df['Prix (TTC)'].sum()]
    })
    return totals_df, glass_total, bottle_total

# Génère un tableau des totaux pour les descriptions sélectionnées
def generate_table_of_totals(dataframe, description_list):
    filtered_df = dataframe[dataframe['Description'].isin(description_list)]
    df = pd.DataFrame({'Description': description_list, 'Quantité': 0, 'Total (en €)': 0})
    for index, row in df.iterrows():
        quantity = filtered_df[filtered_df['Description'] == row['Description']]['Quantité'].sum()
        df.at[index, 'Quantité'] = quantity
        total = filtered_df[filtered_df['Description'] == row['Description']]['Prix (TTC)'].sum()
        df.at[index, 'Total (en €)'] = total
    global_total = df['Total (en €)'].sum().astype(float)
    return df, global_total

# Formate une colonne pour l'affichage avec deux décimales
def display_data(dataframe, column):
    dataframe[column] = dataframe[column].astype(float).map('{:,.2f}'.format)
    return dataframe

# Calcule les ventes par tranches de 30 minutes
def sales_by_half_hours(dataframe):
    grouped_data = dataframe.groupby([pd.Grouper(key='FullDate', freq='30Min')])['Prix (TTC)'].sum().reset_index()
    grouped_data.rename(columns={'Prix (TTC)': 'Total (en €)'}, inplace=True)
    grouped_data['Heure'] = grouped_data['FullDate'].dt.strftime('%H:%M')
    return grouped_data[['Heure', 'Total (en €)']]

# Nettoie les données en les formatant et en ajoutant des colonnes nécessaires
def data_cleaning(data):
    data = data[['Date', 'Quantité', 'Description', 'Prix (TTC)', 'Compte']]
    data = french_to_english(data)
    data['FullDate'] = pd.to_datetime(data['Date'], format='%d %b %Y %H:%M', dayfirst=True)
    data['Heure'] = data['FullDate'].dt.strftime('%H:%M')
    data['Date'] = data['FullDate'].dt.date
    return data

# Formate les colonnes pour Excel en supprimant les décimales inutiles
def formating_for_excel(dataframe, column):
    dataframe[column] = dataframe[column].astype(str).str.replace('.00', '')
    dataframe[column] = dataframe[column].astype(float)
    dataframe['Quantité'] = dataframe['Quantité'].astype(int)
    return dataframe

# Génère un fichier Excel avec les tableaux et les KPI
def to_excel(displayed_global, totaux, total, glass_total, bottle_total):
    # Créer un fichier Excel avec tout sur une seule page
    excel_file = BytesIO()
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        # Convertir le fichier Excel en objet Workbook
        workbook = writer.book
        sheet = workbook.create_sheet(title='Résultats')

        # Désactiver le quadrillage de la feuille
        sheet.sheet_view.showGridLines = False

        # Définir les styles
        header_fill = PatternFill(start_color="941651", end_color="941651", fill_type="solid")
        row_fill_gray = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        alignment = Alignment(horizontal="center", vertical="center")

        # Ajouter displayed_global
        for r_idx, row in enumerate(dataframe_to_rows(displayed_global, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # En-tête
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = alignment
                else:  # Lignes
                    if r_idx % 2 == 0:  # Lignes paires
                        cell.fill = row_fill_gray
                    cell.alignment = alignment
                # Appliquer le format monétaire natif pour la colonne "Prix (TTC)"
                if c_idx == 4:  # Colonne "Prix (TTC)"
                    cell.number_format = '_-* #,##0.00\ [$€-fr-FR]_-'

        # Créer un tableau formaté pour displayed_global
        table_global = Table(displayName="TableGlobal", ref=f"A1:D{len(displayed_global) + 1}")
        style_global = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table_global.tableStyleInfo = style_global
        sheet.add_table(table_global)

        # Ajouter totaux
        start_row_totaux = 1
        start_col_totaux = 7
        for r_idx, row in enumerate(dataframe_to_rows(totaux, index=False, header=True), start=start_row_totaux):
            for c_idx, value in enumerate(row, start=start_col_totaux):
                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # En-tête
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = alignment
                else:  # Lignes
                    if r_idx % 2 == 0:  # Lignes paires
                        cell.fill = row_fill_gray
                    cell.alignment = alignment
                # Appliquer le format monétaire natif pour la colonne "Total (en €)"
                if c_idx == 9:  # Colonne "Total (en €)"
                    cell.number_format = '_-* #,##0.00\ [$€-fr-FR]_-'

        # Créer un tableau formaté pour totaux
        table_totaux = Table(displayName="TableTotaux", ref=f"G1:I{len(totaux) + 1}")
        style_totaux = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table_totaux.tableStyleInfo = style_totaux
        sheet.add_table(table_totaux)

        # Ajouter les KPI
        kpi_data = pd.DataFrame({'KPI': ['Verres vendus', 'Bouteilles vendues'], 'Quantité': [glass_total, bottle_total]})
        start_row_kpi = 1
        start_col_kpi = 12
        for r_idx, row in enumerate(dataframe_to_rows(kpi_data, index=False, header=True), start=start_row_kpi):
            for c_idx, value in enumerate(row, start=start_col_kpi):
                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # En-tête
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = alignment
                else:  # Lignes
                    if r_idx % 2 == 0:  # Lignes paires
                        cell.fill = row_fill_gray
                    cell.alignment = alignment
                # Appliquer le format nombre pour la colonne "Quantité"
                if c_idx == 13:  # Colonne "Quantité"
                    cell.number_format = '0'
                    cell.alignment = alignment

        # Créer un tableau formaté pour les KPIs
        table_kpi = Table(displayName="TableKPI", ref=f"L1:M{len(kpi_data) + 1}")
        style_kpi = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table_kpi.tableStyleInfo = style_kpi
        sheet.add_table(table_kpi)

        # Ajouter le total
        total_label_cell = sheet.cell(row=len(kpi_data) + 3, column=12, value="Montant total")
        total_label_cell.font = Font(bold=True, size=16)
        total_label_cell.alignment = Alignment(horizontal="left", vertical="center")

        total_value_cell = sheet.cell(row=len(kpi_data) + 3, column=13, value=total)
        total_value_cell.font = Font(bold=True, size=16)
        total_value_cell.number_format = '_-* #,##0.00\ [$€-fr-FR]_-'
        total_value_cell.alignment = Alignment(horizontal="right", vertical="center")

    excel_file.seek(0)
    return excel_file

# Application Streamlit
st.set_page_config(
    page_title="Résultats PBLV", 
    page_icon=":wine:", 
    layout="wide")

st.title(':grey[RÉSULTATS PBLV]')
col1, col2,col3 = st.columns([2,2,1])

with st.sidebar:
    sales_data = st.file_uploader('Fichier de ventes', type='.csv')
    if sales_data is not None:
        sales_data = pd.read_csv(sales_data)
    else:
        st.info('Veuillez charger un fichier de ventes',icon='⚠️')
        st.stop()
    st.header('Filtres')
    data = data_cleaning(sales_data)



    #sélecteur de date de départ : la min value et max value sont les dates de début et de fin du fichier de ventes
    first_date = data['Date'].min()
    last_date = data['Date'].max()
    start_date = st.date_input('Date de début', value=None, min_value=first_date, max_value=last_date, key=None)
    #sélecteur de date de fin :
    end_date = st.date_input('Date de fin', value=None, min_value=start_date, max_value=last_date, key=None)
    #data = data[(data['Date'] >= start_date) & (data['Date'] <= end_date)]
    selected_description = st.multiselect('Items concernés', data['Description'].unique())
    st.header('Info Complémentaire')
    last_week_value = st.number_input('Montant de la semaine dernière en euros',format='%.2f', step=.5,min_value=0.0)
    last_week_glass = st.number_input('Verres vendus la semaine dernière',format='%.2f', step=.5,min_value=0.0)
    last_week_bottle = st.number_input('Bouteilles vendues la semaine dernière',format='%.2f', step=.5,min_value=0.0)
    if len(selected_description)==0:
        st.stop()



if selected_description:
    data = filter_data(data, start_date, end_date, selected_description)
    with col1:
        st.header('Données globales de PBLV',divider="red")
        displayed_global = data[['Date', 'Quantité', 'Description', 'Prix (TTC)']]
        st.dataframe(display_data(displayed_global,'Prix (TTC)'),width=500, height=450)

    with col2:
        st.header('Totaux',divider="grey")
        totaux, total = generate_table_of_totals(data, selected_description)
        st.dataframe(display_data(totaux,'Total (en €)'),width=500)
        Verre_description_list = st.multiselect('quelle(s) catégorie(s) correspond(ent) aux **:red[verres]** ?', totaux['Description'].unique())
        Bouteille_description_list = st.multiselect('quelle(s) catégorie(s) correspond(ent) aux **:red[bouteilles]** ?', totaux['Description'].unique())
        hours_totals = sales_by_half_hours(data)
        #st.dataframe(display_data(hours_totals,'Total (en €)'),width=500,  height=200)

        


    with col3:
        st.header('Valeurs clefs',divider="red")
        tot_str, delta_str = metric_in_euros(total,last_week_value)
        st.metric(label="Montant total", value=tot_str,delta=delta_str, border=True)

        if (Verre_description_list is not None) and (Bouteille_description_list is not None):
            Totals_df, glass_total, bottle_total = glass_bottle_total(data, Verre_description_list, Bouteille_description_list)
            st.metric(label="Verres vendus", value=glass_total,delta=glass_total-last_week_glass, border=True)
            st.metric(label="Bouteilles vendues", value=bottle_total,delta=bottle_total-last_week_bottle, border=True)
            

    
#line chart of the total sales by half hours
hours_totals['Total (en €)'] = hours_totals['Total (en €)'].astype(float)

chart = alt.Chart(hours_totals).mark_bar().encode(
    x=alt.X('Heure',axis = alt.Axis(title='Heure', labelAngle=0)),
    y=alt.Y('Total (en €)', axis=alt.Axis(title=None)),
    color=alt.Color('Total (en €)', scale=alt.Scale(range=["#58D68D", "#117A65"]))
)
st.altair_chart(chart, use_container_width=True)




# Ajouter un bouton de téléchargement dans Streamlit
if selected_description:
    # Générer le fichier Excel
    excel_data = to_excel(formating_for_excel(displayed_global,'Prix (TTC)'), formating_for_excel(totaux,'Total (en €)'), total, glass_total, bottle_total)
    
    # Bouton pour télécharger le fichier Excel
    st.download_button(
        label="Télécharger les résultats en Excel",
        data=excel_data,
        file_name="resultats_pblv.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )